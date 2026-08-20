#!/usr/bin/env python3
"""收集表填表数据管理（全局一份 + 每个项目独立一份）。

约定:
  - 每个项目的 21 字段事实源是 <project>/collection.json。
  - sync 之后生成两份 xlsx：
      1) <root>/_shared/收集表_汇总.xlsx   全局汇总（一条一行）
      2) <project>/收集表_<project>.xlsx    该项目独立单行表

用法:
  collection_table.py new --root <dir> --project <name> [--date YYYY-MM-DD]
  collection_table.py write --root <dir> --project <name> --json <file> [--date YYYY-MM-DD]
  collection_table.py sync --root <dir>
  collection_table.py list --root <dir>
"""
import argparse
import json
import os
import re
import sys
from datetime import date as _date
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("缺少 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# (json key, 表头) —— 与甲方 2026-08 新收集表 21 列顺序一致。
FIELDS = [
    ("sample_id", "sample_id"),
    ("session_id", "session id"),
    ("bug_id", "bug_id"),
    ("task_type", "task_type"),
    ("bug_category", "bug_category"),
    ("repo_url", "repo_url"),
    ("go_version", "go_version"),
    ("repro_determinism", "repro_determinism"),
    ("user_query", "user_query"),
    ("trajectory", "trajectory"),
    ("verify_cmds", "verify_cmds"),
    ("gold_root_cause", "gold_root_cause"),
    ("success_criteria", "success_criteria"),
    ("verify_result", "verify_result"),
    ("harness", "harness"),
    ("generator_model", "generator_model"),
    ("worker", "做题人"),
    ("creator", "创建人"),
    ("qc_result", "质检结果"),
    ("qc_note", "质检备注"),
    ("sync_feishu", "是否同步飞书"),
]
HEADERS = [h for _, h in FIELDS]
WRAP_COLS = {"repo_url", "user_query", "verify_cmds", "gold_root_cause", "success_criteria", "verify_result"}

# user_query 是给测试模型的自然语言求助，不得夹带验收动作或具体运行命令。
# 公开行为预期（如“取消后应停止”）仍然允许；这里只拦截要求模型执行验证的措辞。
USER_QUERY_ACCEPTANCE_PATTERNS = [
    r"\bgo\s+(?:test|run|build|vet)\b",
    r"\b(?:pytest|npm\s+(?:test|run)|make\s+test)\b",
    r"(?:请|麻烦|帮我|你)?\s*(?:运行|执行|跑|跑一下|跑下|重跑|复现)\s*(?:这个|该|相关|定向)?\s*(?:测试|用例|命令|脚本)",
    r"(?:通过|用|执行)\s*(?:测试|命令)\s*(?:确认|验证|检查)",
    r"(?:验收|验证|复现)\s*(?:一下|结果|是否|这个问题|该问题)",
    r"(?:修复后|改完后|完成后)\s*(?:请|要)?\s*(?:运行|执行|跑|验证|检查)",
]


def user_query_acceptance_issues(text: str) -> list[str]:
    """返回 user_query 中夹带的验收/复现指令；空列表表示通过。"""
    value = str(text or "").strip()
    if not value:
        return []
    return [pattern for pattern in USER_QUERY_ACCEPTANCE_PATTERNS
            if re.search(pattern, value, flags=re.IGNORECASE)]


def date_dir(root: Path, date: str) -> Path:
    return root / (date or _date.today().isoformat())


def find_project(root: Path, name: str, date: str | None) -> Path | None:
    name = name.replace("/", "__").lower()
    dirs = [date_dir(root, date)] if date else sorted(root.glob("*/"))
    for d in dirs:
        d = Path(d)
        if not d.is_dir() or d.name.startswith("_"):
            continue
        for child in d.iterdir():
            if child.is_dir() and child.name.lower() == name:
                return child
    return None


# verify_result 现为机器生成的 pre_fix/post_fix JSON，不参与文案雷同检查
_DUP_FIELDS = ["user_query", "success_criteria", "verify_cmds"]


def _dup_issues_for(root: Path, data: dict, project_dir: Path):
    """把即将写入的一条记录，与 root 下已存在的记录做文案雷同比对。

    返回 (banned_issues, dup_issues, forbidden_issues)，三条都只统计涉及本记录的硬红。
    复用 check_prompt_duplicates.py 的判定逻辑，保证口径一致。
    """
    from check_prompt_duplicates import (
        check_banned as _check_banned,
        check_forbidden_chars as _check_forbidden,
        collect as _collect,
        find_duplicates as _find_dups,
    )
    project_id = project_dir.name
    target = project_dir.resolve()
    existing = []
    for r in _collect(root):
        try:
            if Path(r["dir"]).resolve() != target:
                existing.append(r)
        except Exception:
            existing.append(r)
    incoming_texts = [
        {"field": f, "text": str(data.get(f) or "").strip(), "source": "write", "path": "<incoming>"}
        for f in _DUP_FIELDS if str(data.get(f) or "").strip()
    ]
    incoming = {"id": project_id, "dir": "<incoming>", "texts": incoming_texts}
    records = existing + [incoming]

    banned = [b for b in _check_banned(records) if b["id"] == project_id]
    dups = [d for d in _find_dups(records, 12) if project_id in d["ids"]]
    forbidden = [x for x in _check_forbidden(records) if x["id"] == project_id]
    return banned, dups, forbidden


def style_sheet(ws):
    ws.row_dimensions[1].height = 24
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 28
    header_key_by_idx = {i + 1: FIELDS[i][0] for i in range(len(FIELDS))}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True) \
                if header_key_by_idx.get(cell.column) in WRAP_COLS \
                else Alignment(vertical="top")


def write_xlsx(path: Path, rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(key, "") for key, _ in FIELDS])
    style_sheet(ws)
    wb.save(path)
    return path


def collect_rows(root: Path) -> list[tuple[Path, dict]]:
    out = []
    for d in sorted(root.glob("*/")):
        if d.name.startswith("_"):
            continue
        for child in sorted(d.iterdir()):
            f = child / "collection.json"
            if f.exists():
                try:
                    out.append((child, json.loads(f.read_text(encoding="utf-8"))))
                except json.JSONDecodeError as e:
                    print(f"⚠️  跳过损坏的 {f}: {e}")
    return out


def cmd_new(args):
    root = Path(args.root)
    p = find_project(root, args.project, args.date)
    if not p:
        print(f"❌ 找不到项目: {args.project}（请先 workspace.py new-project）")
        sys.exit(1)
    f = p / "collection.json"
    if f.exists():
        print(f"⚠️  已存在 {f}，不覆盖。用 write 更新。")
        sys.exit(1)
    template = {key: "" for key, _ in FIELDS}
    template["bug_id"] = f"{p.name.lower()}-0001"
    f.write_text(json.dumps(template, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"✅ 已生成 {f}")


def cmd_write(args):
    root = Path(args.root)
    p = find_project(root, args.project, args.date)
    if not p:
        print(f"❌ 找不到项目: {args.project}")
        sys.exit(1)
    data = json.loads(Path(args.json).read_text(encoding="utf-8"))
    known = {key for key, _ in FIELDS}
    unknown = set(data) - known
    if unknown:
        print(f"⚠️  存在未知字段（会被忽略）: {sorted(unknown)}")
    clean = {key: data.get(key, "") for key, _ in FIELDS}
    # 是否同步飞书由甲方/录入方决定，本脚本默认不自动填写。
    if clean.get("sync_feishu") is None:
        clean["sync_feishu"] = ""
    task_type = str(clean.get("task_type") or "").strip()
    verify_command = str(clean.get("verify_cmds") or "").strip()
    user_query = str(clean.get("user_query") or "").strip()
    prompt_file = p / "prompt.txt"
    prompt_text = prompt_file.read_text(encoding="utf-8").strip() if prompt_file.exists() else ""
    query_sources = [("collection.json", user_query)]
    if prompt_text:
        query_sources.append(("prompt.txt", prompt_text))
    acceptance_issues = [
        (source, issues)
        for source, text in query_sources
        if (issues := user_query_acceptance_issues(text))
    ]
    if acceptance_issues:
        print("❌ user_query 纯提示词硬门禁：不得写验收/复现/运行指令；请删除具体命令及要求模型运行、验证或复现测试的句子。")
        for source, patterns in acceptance_issues:
            for pattern in patterns:
                print(f"   [{source}] [matched] {pattern}")
        sys.exit(1)
    harness = str(clean.get("harness") or "").strip()
    if harness and not re.search(r"\bv?\d+(?:\.\d+)+\b", harness):
        print("❌ harness 必须写明生成轨迹的工具名 + 版本号，例如 Claude Code CLI v2.1.233")
        sys.exit(1)
    if task_type in ("bugfix", "diagnosis"):
        from verify_cmds import (
            CONCURRENCY_CATEGORY,
            validate_concurrency_metadata,
            validate_delivery_field_wording,
            validate_success_criteria,
            validate_verify_cmds,
        )
        require_race = str(clean.get("bug_category") or "").strip() == CONCURRENCY_CATEGORY
        verify_issues = validate_verify_cmds(verify_command, require_race=require_race)
        verify_issues.extend(validate_concurrency_metadata(clean))
        verify_issues.extend(validate_success_criteria(clean))
        verify_issues.extend(validate_delivery_field_wording(clean))
        if prompt_text and prompt_text != user_query:
            verify_issues.extend(validate_delivery_field_wording({"user_query": prompt_text}))
        if verify_issues:
            print("❌ 收集表硬门禁：" + "；".join(verify_issues))
            sys.exit(1)
    elif task_type:
        print("❌ task_type 只能是 bugfix 或 diagnosis")
        sys.exit(1)
    if not args.skip_dup_check:
        banned, dups, forbidden = _dup_issues_for(root, clean, p)
        if banned or dups or forbidden:
            print("❌ 文案硬门禁：本条命中「去 AI 味 / 去同质化 / 生僻字符」红线，拒绝写入。请改写后重试；如确认要强制写入，加 --skip-dup-check。")
            for b in banned:
                print(f"   [banned] {b['field']}: {b['phrase']!r}")
            for x in forbidden:
                print(f"   [forbidden] {x['field']}: 含生僻/序号字符 {x['chars']!r}")
            for d in dups:
                others = sorted(set(d["ids"]) - {p.name})
                print(f"   [dup] {d['field']}: {d['sub']!r} 与 {others} 重复")
            sys.exit(1)
    (p / "collection.json").write_text(json.dumps(clean, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_xlsx(p / f"收集表_{p.name}.xlsx", [clean])
    cmd_sync(args)
    print(f"✅ 已写入 {p / 'collection.json'}")
    print(f"✅ 项目独立表: {p / f'收集表_{p.name}.xlsx'}")


def cmd_sync(args):
    root = Path(args.root)
    rows = collect_rows(root)
    if not rows:
        print("（没有任何 collection.json，跳过汇总）")
        return
    write_xlsx(root / "_shared" / "收集表_汇总.xlsx", [r for _, r in rows])
    print(f"✅ 全局汇总: {root / '_shared' / '收集表_汇总.xlsx'}（{len(rows)} 行）")
    for p, row in rows:
        f = p / f"收集表_{p.name}.xlsx"
        write_xlsx(f, [row])
        print(f"   - {f}")


def cmd_list(args):
    root = Path(args.root)
    rows = collect_rows(root)
    if not rows:
        print("（没有项目填写 collection.json）")
        return
    print(f"{'日期':12} {'项目':40} sample_id  bug_id")
    print("-" * 90)
    for p, row in rows:
        print(f"{p.parent.name:12} {p.name:40} {row.get('sample_id','')}  {row.get('bug_id','')}")


def main():
    p = argparse.ArgumentParser(description="收集表填表数据管理")
    sub = p.add_subparsers(dest="cmd", required=True)

    c = sub.add_parser("new")
    c.add_argument("--root", default=".")
    c.add_argument("--project", required=True)
    c.add_argument("--date")
    c.set_defaults(func=cmd_new)

    c = sub.add_parser("write")
    c.add_argument("--root", default=".")
    c.add_argument("--project", required=True)
    c.add_argument("--json", required=True)
    c.add_argument("--date")
    c.add_argument("--skip-dup-check", action="store_true", help="跳过文案雷同硬门禁（不推荐）")
    c.set_defaults(func=cmd_write)

    c = sub.add_parser("sync")
    c.add_argument("--root", default=".")
    c.set_defaults(func=cmd_sync)

    c = sub.add_parser("list")
    c.add_argument("--root", default=".")
    c.set_defaults(func=cmd_list)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
